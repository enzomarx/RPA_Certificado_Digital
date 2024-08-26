import os
import time
import yagmail
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC
import schedule
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Configurações do Selenium
download_dir = os.path.expanduser("~/Downloads")  # Diretório de downloads do PC
chrome_options = Options()
chrome_options.add_experimental_option("prefs", {
    "download.default_directory": r"C:\Users\PC\Downloads",
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True,
    "profile.default_content_settings.popups": 0
})

# Inicializando o WebDriver
driver = webdriver.Chrome(service=Service(), options=chrome_options)

# Função para encontrar o arquivo mais recente
def find_latest_file():
    files = [f for f in os.listdir(download_dir) if f.startswith("Portal") and f.endswith(".xlsx")]
    files.sort(key=lambda x: os.path.getmtime(os.path.join(download_dir, x)), reverse=True)
    return files[0] if files else None

# Função para verificar se o arquivo foi completamente baixado
def wait_for_download(filename, timeout=60):
    start_time = time.time()
    file_path = os.path.join(download_dir, filename)
    while time.time() - start_time < timeout:
        if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
            if not filename.endswith('.crdownload'):
                return file_path
        time.sleep(1)
    return None

# Função para modificar o arquivo Excel
def modify_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Desmesclar a célula A1:D1
    ws.unmerge_cells('A1:D1')
    
    # Modificar o texto da célula A1 e aplicar negrito
    ws['A1'] = "Relatório Certificado Digital"
    ws['A1'].font = Font(bold=True)
    
    # Adicionar o texto "Senha" na célula D2 e aplicar negrito
    ws['D2'] = "Senha"
    ws['D2'].font = Font(bold=True)
    
    # Adicionar o texto "Vencimento em" na célula E2 e aplicar negrito
    ws['E2'] = "Vencimento em"
    ws['E2'].font = Font(bold=True)
    
    # Adicionar a fórmula para calcular os dias até o vencimento na coluna E
    for row in range(3, ws.max_row + 1):  # Começa da linha 3, considerando que a linha 2 tem títulos
        ws[f'E{row}'] = f"=C{row}-TODAY()"
        # Aplicar cor vermelha ao texto na coluna E
        ws[f'E{row}'].font = Font(color="FF0000")
    
    # Aplicar cor vermelha ao texto na célula E2
    ws['E2'].font = Font(bold=True, color="FF0000")
    
    # Centralizar o conteúdo da coluna B, com exceção da célula B2
    for row in range(3, ws.max_row + 1):  # Começa da linha 3
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
    
    # Alinhar à direita o conteúdo da coluna D, com exceção da célula D2
    for row in range(3, ws.max_row + 1):  # Começa da linha 3
        ws[f'D{row}'].alignment = Alignment(horizontal='right')
    
    # Salvar as modificações
    wb.save(file_path)
    print(f"Modificações feitas no arquivo {file_path}")

# Função de enviar e-mails
def send_email():
    recipients = ["cassandra@controllersbr.com", "ti2.controllersbr@gmail.com", "carlos.junior@controllersbr.com", "joas@controllersbr.com", "lucas@controllersbr.com", "juliocesar@controllersbr.com"]
    sender_email = "redstarenzo@gmail.com"
    yag = yagmail.SMTP(sender_email, "evms sijx toii aqxx")
    
    # Encontrar o arquivo mais recente
    filename = find_latest_file()
    if filename:
        # Esperar o arquivo ser completamente baixado
        file_path = wait_for_download(os.path.basename(filename))
        if file_path:
            # Modificar o arquivo Excel
            modify_excel(file_path)
            
            # Enviar o e-mail com o arquivo modificado
            yag.send(
                to=recipients,
                subject="Relatório Automático (Certificado Digital)",  # Ajuste conforme necessário
                contents="Por favor, veja a planilha em anexo.",
                attachments=file_path,
            )
            print(f"E-mail enviado com sucesso com o anexo {file_path}")
        else:
            print("Arquivo não encontrado ou não está completo.")
    else:
        print("Nenhum arquivo encontrado para enviar.")
        
# Função para executar a automação
def run_automation():
    # Setup
    driver.get("https://08978175000180.portal-veri.com.br")
    driver.set_window_size(1366, 768)

    # Ações no Login
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="kt_sign_in_form"]/div[2]/input'))).send_keys("ti2.controllersbr@gmail.com")  # Ajuste conforme necessário
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="kt_sign_in_form"]/div[3]/input'))).send_keys("senha123")  # Ajuste conforme necessário
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="kt_sign_in_form"]/div[5]/button'))).click()

    # Ações no Dashboard
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".row:nth-child(1) > .col-md-3:nth-child(3) .ms-3"))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".col-md-3:nth-child(4) a:nth-child(1)"))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".buttons-excel > span"))).click()
    
    # Aguardar download ser concluído
    time.sleep(10)
    
    # Enviar o e-mail já com documento baixado e modificado
    send_email()
    time.sleep(10)
    
    # Fechar o navegador
    driver.quit()

# Agendamento
#def job():
run_automation()

#schedule.every().monday.at("09:00").do(job)

# Mantém o script rodando para cumprir agendamento
#while True:
#    schedule.run_pending()
#    time.sleep(1)
